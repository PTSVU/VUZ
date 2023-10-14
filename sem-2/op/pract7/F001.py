import vk_api
from vk_api.longpoll import VkLongPoll, VkEventType
from vk_api.utils import get_random_id
import json
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
import re
from bs4 import BeautifulSoup
import requests
import os
import datetime
import openpyxl
from openpyxl import load_workbook
import calendar
import locale


token = 'vk1.a.l4UFE1Ln6Q-QSJdkXkuiFX5fb-xRvJc0MWoJhAugZHG7pOlYkZ29fHygEkZGz7gTMKQjIduwNP9Ur-a7nLF6ZFmh3nVcb5Dw6wF__97heKRungaskXbCvNmyuiDt_sVfZ_0ROTdmlXwjzbDv-hxKIDLgZm5OLAwP7M2KWQS_HkaA5bsV4ror16qDYE-YsLni4P4bCPf79kkfdPAupmRPcQ'


locale.setlocale(locale.LC_ALL, 'ru_RU.utf8')
calendar.different_locale('ru_RU.utf8')


def save_group(id, group):
    with open("users_group.json", "r") as file:
        data = json.load(file)
    data[str(id)] = str(group)
    with open("users_group.json", "w") as file:
        json.dump(data, file)


def load_group(id):
    try:
        with open("users_group.json", "r") as file:
            data = json.load(file)
        return data[str(id)]
    except KeyError:
        return None


def get_xlsx():
    download_directory = "temp_d"
    os.makedirs(download_directory, exist_ok=True)
    files = os.listdir(download_directory)
    for file_name in files:
        file_path = os.path.join(download_directory, file_name)
        os.remove(file_path)
    url = "https://www.mirea.ru/schedule/"
    page = requests.get(url)
    soup = BeautifulSoup(page.text, "html.parser")
    result = soup.find("div", class_="schedule"). \
        find(string="Институт информационных технологий"). \
        find_parent("div"). \
        find_parent("div"). \
        findAll("a", class_="uk-link-toggle")
    filename = 0
    for link in result:
        href = link.get('href')
        if href is not None:
            filename += 1
            file_path = os.path.join(download_directory, "kurs_"+(str(filename)+"_IIT"+".xlsx"))
            with open(file_path, "wb") as file:
                resp = requests.get(href)
                file.write(resp.content)

def week_number():
    today = datetime.date.today()
    week_number = today.isocalendar()[1]
    return week_number - 5


def schedule(group):
    sch_a = []
    sch_b = []

    if group[-2:] == "22":
        workbook = openpyxl.load_workbook('temp_d/kurs_1_IIT.xlsx')
    elif group[-2:] == "21":
        workbook = openpyxl.load_workbook('temp_d/kurs_2_IIT.xlsx')
    elif group[-2:] == "20":
        workbook = openpyxl.load_workbook('temp_d/kurs_3_IIT.xlsx')
    else:
        print("Такой группы не найдено")
        return None
    sheet = workbook.active
    row_values = [cell.value for cell in sheet[2]]
    target_text = group
    for idx, value in enumerate(row_values):
        if value == target_text:
            column = idx + 1
            break
    else:
        print("Такой группы не найдено")
        workbook.close()
        return None
    for row in range(4, sheet.max_row + 1):
        par = sheet.cell(row=row, column=column).value
        if par == "":
            par = "-"
        par_type = sheet.cell(row=row, column=column+1).value
        par_pred = sheet.cell(row=row, column=column+2).value
        if par_pred == "":
            par_pred = "-"
        par_aud = sheet.cell(row=row, column=column+3).value
        if sheet.cell(row=row, column=5).value == "I":
            if row == 4 or row == 18 or row == 32 or row == 46 or row == 60 or row == 74:
                sch_a.append(sheet.cell(row=row, column=1).value)
            par_num = sheet.cell(row=row, column=2).value
            if par == "-":
                sch_a.append((str(par_num) + ") " + str(par)))
            else:
                sch_a.append((str(par_num) + ") " + str(par) + ", " + str(par_type) + ", " + str(par_pred)
                          + ", " + str(par_aud)).replace('\n', ''))
        elif sheet.cell(row=row, column=5).value == "II":
            if row-1 == 4 or row-1 == 18 or row-1 == 32 or row-1 == 46 or row-1 == 60 or row-1 == 74:
                sch_b.append(sheet.cell(row=row-1, column=1).value)
            par_num = sheet.cell(row=row-1, column=2).value
            if par == "-":
                sch_b.append((str(par_num) + ") " + str(par)))
            else:
                sch_b.append((str(par_num) + ") " + str(par) + ", " + str(par_type) + ", " + str(par_pred)
                          + ", " + str(par_aud)).replace('\n', ''))
        if sheet.cell(row=row+1, column=2).value == "":
            break
    workbook.close()
    return sch_a, sch_b


def main():
    get_xlsx()
    vk_session = vk_api.VkApi(token=token)
    vk = vk_session.get_api()
    longpoll = VkLongPoll(vk_session)

    keyboard = VkKeyboard(one_time=True)
    keyboard.add_button('на сегодня', color=VkKeyboardColor.POSITIVE)
    keyboard.add_button('на завтра', color=VkKeyboardColor.NEGATIVE)
    keyboard.add_line()
    keyboard.add_button('на эту неделю', color=VkKeyboardColor.PRIMARY)
    keyboard.add_button('на следующую неделю', color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button('какая неделя?', color=VkKeyboardColor.SECONDARY)
    keyboard.add_button('какая группа?', color=VkKeyboardColor.SECONDARY)


    for event in longpoll.listen():
        if event.type == VkEventType.MESSAGE_NEW and event.to_me and event.text:
            print('id{}: "{}"'.format(event.user_id, event.text), end='\n')
            if event.text.lower() == "начать":
                vk.messages.send(
                    user_id=event.user_id,
                    random_id=get_random_id(),
                    message=f'Привет, {vk.users.get(user_id = event.user_id)[0]["first_name"]}\n'
                            f'Этот бот умеет показывать рассписание и проду\n'
                            f'Что бы продолжить напиши "Бот"'
                )
            elif event.text.lower() == "бот":
                vk.messages.send(
                    user_id=event.user_id,
                    random_id=get_random_id(),
                    keyboard=keyboard.get_keyboard(),
                    message=f'Показать расписание …'
                )
            elif re.search("\A\w{4}-\d{2}-\d{2}\Z", event.text):
                temp_text = str(re.findall(r"\w{4}-\d{2}-\d{2}", event.text.upper())).replace("[", "").replace("]", "").replace("'", "")
                vk.messages.send(
                    user_id=event.user_id,
                    random_id=get_random_id(),
                    keyboard=keyboard.get_keyboard(),
                    message='Я запомнил, что ты из группы ' + temp_text
                )
                save_group(event.user_id, temp_text)
            elif event.text.lower() == "на сегодня":
                today = calendar.day_name[datetime.datetime.today().day % 7 - 1]
                tmr = calendar.day_name[datetime.datetime.today().day % 7]
                ch = ""
                nch = ""
                sch_a, sch_b = schedule(load_group(event.user_id))
                if week_number() % 2 == 0:
                    td = 0
                    for i in range(len(sch_b)):
                        if sch_b[i].lower() == today:
                            td = 1
                        if sch_b[i].lower() == tmr:
                            td = 0
                        if td == 1:
                            ch += sch_b[i]
                            if i != len(sch_b):
                                ch += "\n"
                    vk.messages.send(
                        user_id=event.user_id,
                        random_id=get_random_id(),
                        keyboard=keyboard.get_keyboard(),
                        message=f"{ch}"
                    )
                elif event.text.lower() == "на сегодня":
                    today = calendar.day_name[datetime.datetime.today().day % 7 - 1]
                    tmr = calendar.day_name[datetime.datetime.today().day % 7]
                    ch = ""
                    nch = ""
                    sch_a, sch_b = schedule(load_group(event.user_id))
                    if week_number() % 2 == 0:
                        td = 0
                        for i in range(len(sch_b)):
                            if sch_b[i].lower() == today:
                                td = 1
                            if sch_b[i].lower() == tmr:
                                td = 0
                            if td == 1:
                                ch += sch_b[i]
                                if i != len(sch_b):
                                    ch += "\n"
                        vk.messages.send(
                            user_id=event.user_id,
                            random_id=get_random_id(),
                            keyboard=keyboard.get_keyboard(),
                            message=f"{ch}"
                        )
                elif week_number() % 2 != 0:
                    td = 0
                    for i in range(len(sch_a)):
                        if sch_a[i].lower() == today:
                            td = 1
                        if sch_a[i].lower() == tmr:
                            td = 0
                        if td == 1:
                            nch += sch_a[i]
                            if i != len(sch_a):
                                nch += "\n"
                    vk.messages.send(
                        user_id=event.user_id,
                        random_id=get_random_id(),
                        keyboard=keyboard.get_keyboard(),
                        message=f"{nch}"
                    )
            elif event.text.lower() == "на завтра":
                today = calendar.day_name[datetime.datetime.today().day % 7]
                tmr = calendar.day_name[datetime.datetime.today().day % 7 + 1]
                ch = ""
                nch = ""
                sch_a, sch_b = schedule(load_group(event.user_id))
                if week_number() % 2 == 0:
                    td = 0
                    for i in range(len(sch_b)):
                        if sch_b[i].lower() == today:
                            td = 1
                        if sch_b[i].lower() == tmr:
                            td = 0
                        if td == 1:
                            ch += sch_b[i]
                            if i != len(sch_b):
                                ch += "\n"
                    vk.messages.send(
                        user_id=event.user_id,
                        random_id=get_random_id(),
                        keyboard=keyboard.get_keyboard(),
                        message=f"{ch}"
                    )
                elif week_number() % 2 != 0:
                    td = 0
                    for i in range(len(sch_a)):
                        if sch_a[i].lower() == today:
                            td = 1
                        if sch_a[i].lower() == tmr:
                            td = 0
                        if td == 1:
                            nch += sch_a[i]
                            if i != len(sch_a):
                                nch += "\n"
                    vk.messages.send(
                        user_id=event.user_id,
                        random_id=get_random_id(),
                        keyboard=keyboard.get_keyboard(),
                        message=f"{nch}"
                    )
            elif event.text.lower() == "на завтра":
                pass
            elif event.text.lower() == "на эту неделю":
                pass
            elif event.text.lower() == "на следующую неделю":
                pass
            elif event.text.lower() == "какая неделя?":
                vk.messages.send(
                    user_id=event.user_id,
                    random_id=get_random_id(),
                    keyboard=keyboard.get_keyboard(),
                    message=f"Идёт {week_number()} неделя"
                )
            elif event.text.lower() == "какая группа?":
                vk.messages.send(
                    user_id=event.user_id,
                    random_id=get_random_id(),
                    keyboard=keyboard.get_keyboard(),
                    message=load_group(event.user_id)
                )
            else:
                vk.messages.send(
                    user_id=event.user_id,
                    random_id=get_random_id(),
                    message='Неизветная комманда'
                )


if __name__ == "__main__":
    main()