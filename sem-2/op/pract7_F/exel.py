from openpyxl import load_workbook
import calendar
import os
import time
from datetime import date

from openpyxl import load_workbook

from schedule import *


def delete_excel_file_after_delay(file_path, delay_hours):
    # Ожидание указанного количества часов
    time.sleep(delay_hours * 3600)

    # Удаление файла
    try:
        os.remove(file_path)
        print(f"Файл {file_path} успешно удален.")
    except OSError as e:
        print(f"Ошибка при удалении файла {file_path}: {e}")



delay_hours = 2


globalNumber = -1
isSetGlobalNumber = False


def search_excel(file_path, search_value):
    global globalNumber, isSetGlobalNumber
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Get the second row
    second_row = list(sheet.iter_rows(min_row=2, max_row=2))

    # Search the second row for the search_value
    for cell in second_row[0]:
        if cell.value == search_value:
            # Match found
            column_number = cell.column
            globalNumber = column_number

            isSetGlobalNumber = True
            return column_number

    # Match not found
    print("Match not found")
    return None


def is_even(number):
    if number % 2 == 0:
        return True
    else:
        return False


def get_day_schedule(file_path, day, week):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    column_number = globalNumber

    row_counter = 0
    lesson_number = 1
    day_schedule = []
    if column_number is None:
        return None

    if day == "Monday":
        min_row = 4
        max_row = 17
    elif day == "Tuesday":
        min_row = 18
        max_row = 31
    elif day == "Wednesday":
        min_row = 32
        max_row = 45
    elif day == "Thursday":
        min_row = 46
        max_row = 59
    elif day == "Friday":
        min_row = 60
        max_row = 73
    elif day == "Saturday":
        min_row = 74
        max_row = 87
    elif day == 'Sunday':
        return ["ЗаНяТиЙ в ВоСкРеСенЬе нЕт!"]

    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=column_number, max_col=column_number + 3):
        lesson_info = []
        for cell in row:
            if cell.column == column_number:
                if is_even(week) and row_counter % 2 == 0:
                    if cell.value:
                        lesson_info.append(f"{lesson_number}) {cell.value}")
                    else:
                        lesson_info.append(f"{lesson_number}) -")
                    lesson_number += 1
                elif not is_even(week) and row_counter % 2 != 0:
                    if cell.value:
                        lesson_info.append(f"{lesson_number}) {cell.value}")
                    else:
                        lesson_info.append(f"{lesson_number}) -")
                    lesson_number += 1
            elif (is_even(week) and row_counter % 2 == 0) or (not is_even(week) and row_counter % 2 != 0):
                if cell.value:
                    lesson_info.append(cell.value)

        if lesson_info:
            day_schedule.append(lesson_info)
        row_counter += 1

    return day_schedule


russian_weekdays = {
    'Monday': 'Понедельник',
    'Tuesday': 'Вторник',
    'Wednesday': 'Среда',
    'Thursday': 'Четверг',
    'Friday': 'Пятница',
    'Saturday': 'Суббота',
}


def get_week_schedule(file_path, week_number):
    week_schedule = {}
    for eng_day, rus_day in russian_weekdays.items():
        schedule = get_day_schedule(file_path, eng_day, week_number)
        if schedule is None:
            return None
        else:
            formatted_schedule = '\n'.join([' '.join(lesson_info) for lesson_info in schedule])
            week_schedule[rus_day] = formatted_schedule
    return week_schedule


def what_day_is_it():
    my_date = date.today()
    return calendar.day_name[my_date.weekday()]


import datetime


def get_current_week_number():
    today = datetime.date.today()
    week_number = today.isocalendar()[1]
    return week_number


import re


def check_group_format(group_number):
    pattern = r'^[A-Za-zА-Яа-я]{4}-\d{2}-\d{2}$'
    match = re.match(pattern, group_number)
    if match:
        return True
    else:
        return False


def capitalize_word(word):
    capitalized_word = word.upper()
    return capitalized_word


nameOfPath = 'IIT_'


def create_file(group_number, links):
    if check_group_format(group_number):
        last_two_digits = int(group_number[-2:])
        if last_two_digits == 22:
            link = links[0]
        elif last_two_digits == 21:
            link = links[1]
        elif last_two_digits == 20:
            link = links[2]
        else:
            return None

        response = requests.get(link)
        file_path = f'{nameOfPath + group_number[-2:]}.xlsx'
        with open(file_path, 'wb') as f:
            f.write(response.content)
        return file_path
    else:
        return None


def check_bot_and_group_format(s):
    words = s.split()
    if len(words) == 2 and words[0].strip == "бот " and check_group_format(words[1].strip()):
        return True
    else:
        return False

def save_user_group(user_id, group):
    with open('user_group.csv', 'a') as f:
        f.write(f"{user_id},{group}\n")

