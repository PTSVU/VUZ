import json
import re
from copy import deepcopy as copy

import openpyxl
import requests
from parsing_links import results

weekdays = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота", "воскресенье"]
group_regex = r"И[АВКНМ]{1}БО-[0-9]{2}-1[7-9]{1}"
pattern = r'^[A-Za-zА-Яа-я]{4}-\d{2}-\d{2}$'

# print(results)
professors = {}


def set_professor(name, lesson, week_day, para_num, week_oddity):
    if name not in professors:
        day = [[None] * 2, [None] * 2, [None] * 2, [None] * 2, [None] * 2, [None] * 2]
        week = {'понедельник': copy(day), 'вторник': copy(day), 'среда': copy(day), 'четверг': copy(day),
                'пятница': copy(day),
                'суббота': copy(day)}
        professors.update({name: week})

    professors[name][week_day][para_num][week_oddity] = lesson


def find_professors_by_last_name(last_name):
    return [key for key, value in professors.items() if key.startswith(last_name)]


def get_course_schedule(course=1, request_res=results):
    course_s = str(course) + "-kurs"
    filename = "schedule_" + course_s + ".xlsx"
    file_regex = r"IIT_" + str(course) + r"-kurs_\d{2}_\d{2}_\w+_\d{2}\.\d{2}\.\d{4}\.xlsx"

    for result in request_res:
        if re.search(file_regex, str(result), re.IGNORECASE) is not None:
            f = open(filename, "wb")
            downloaded_file = requests.get(result["href"])
            f.write(downloaded_file.content)
            f.close()

    book = openpyxl.load_workbook(filename)
    sheet = book.active

    num_cols = sheet.max_column
    num_rows = sheet.max_row

    groups = []
    schedule = {}

    for col_index in range(1, num_cols + 1):
        group_name = str(sheet.cell(row=2, column=col_index).value)
        if re.search(pattern, group_name, re.IGNORECASE) is not None:
            groups.append(group_name)
            week = {"понедельник": None, "вторник": None, "среда": None, "четверг": None, "пятница": None,
                    "суббота": None, "воскресенье": "Выходной! Пар нет"}
            for day in range(6):
                lessons = [[], [], [], [], [], [], []]
                for para in range(7):
                    for week_oddity in range(2):
                        row_index = 4 + week_oddity + (para * 2) + (day * 14)
                        if row_index > 87:
                            break
                        subject = sheet.cell(row_index, col_index).value

                        if subject == '' or subject == None:
                            subject = "--"
                        else:
                            subject = subject.replace("\n", "; ")
                        lesson_type = sheet.cell(row_index, col_index + 1).value
                        if lesson_type == '' or lesson_type == None:
                            lesson_type = "--"
                        else:
                            lesson_type = lesson_type.replace("\n", "; ")

                        lecturer = sheet.cell(row_index, col_index + 2).value
                        if lecturer == '' or lecturer == None:
                            lecturer = "--"
                        else:

                            lecturer = lecturer.replace("\n", "; ")

                        classroom = sheet.cell(row_index, col_index + 3).value
                        if classroom == '' or classroom == None:
                            classroom = "--"
                        else:
                            classroom = classroom.replace("\n", "; ")
                        url = sheet.cell(row_index, col_index + 4).value
                        if url == '' or url == None:
                            url = "--"
                        else:
                            url = url.replace("\n", "; ")
                        lesson = {"Предмет": subject, "Вид занятий": lesson_type, "Преподаватель": lecturer,
                                  "Аудитория": classroom, "Ссылка": url}
                        lessons[para].append(lesson)

                        professors_list = lecturer.split('; ')
                        subject_list = subject.split('; ')
                        pr_lesson = copy(lesson)
                        pr_lesson.pop('Преподаватель')
                        pr_lesson['Группа'] = group_name

                        for h in range(len(professors_list)):
                            if len(subject_list) > h:
                                pr_lesson['Предмет'] = subject_list[h]
                            set_professor(professors_list[h], pr_lesson, weekdays[day], day, week_oddity)

                week[weekdays[day]] = lessons
            schedule.update({group_name: week})
    # print(schedule)
    return schedule


schedule_first = get_course_schedule()
schedule_second = get_course_schedule(2)
schedule_third = get_course_schedule(3)

#print(schedule_first)
#print(professors)
professor_names = list(professors.keys())


def search_professor_by_surname(surname: str, professor_names_clear: list) -> list:
    result = []
    for professor in professor_names_clear:
        if surname == professor.split()[0]:
            result.append(professor)
    return result


# разделить имена преподавателей по запятой
professor_names = [name.strip() for name in ','.join(professor_names).split(',')]

# убрать дубликаты
professor_names = list(set(professor_names))
pattern_for_proffesors = r'\d+\s+[п|П]/[г|Г]'
professor_names_clear = [i for i in professor_names if i not in ['', '--'] and not re.search(pattern_for_proffesors, i)]

#print(professor_names_clear)
#print(search_professor_by_surname("Красников", professor_names_clear))

with open("course1_sch.json", "w") as write_file:
    json.dump(schedule_first, write_file)
with open("course2_sch.json", "w") as write_file:
    json.dump(schedule_second, write_file)
with open("course3_sch.json", "w") as write_file:
    json.dump(schedule_third, write_file)

with open("professors.json", 'w') as write_file:
    write_file.write(json.dumps(professors, ensure_ascii=False, indent=4))
